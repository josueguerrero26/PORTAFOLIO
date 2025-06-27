import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re
import os

# Comproba si existe el archivo de Excel
nombre_archivo = "datos.xlsx"
if os .path.exists(nombre_archivo):
    # Cargar el libro de Excel existente
    wb = load_workbook(nombre_archivo)
    ws = wb.active
else:
# Crear el libro de excel 
    wb = Workbook()
    ws = wb.active
    ws.append(["Nombre", "Apellido", "Edad", "Telefono", "Email"])
    wb.save("datos.xlsx")

def guardar_datos():
    nombre = entry_nombre.get()
    apellido= entry_apellido.get()
    edad = entry_edad.get()
    email = entry_email.get()
    telefono = entry_telefono.get()
    # Validar que los campos no estén vacíos
    if not nombre or not apellido or not edad or not telefono or not email:
        messagebox.showwarning(title="Advertencia", message="Todos los campos son obligatorios")
        return
    try: 
        edad = int(edad)
        telefono = int(telefono)
    except ValueError:
        messagebox.showerror(title="Error", message="Edad y Telefono deben ser números")
        return
    # Validar el formato del email
    if not re.match(pattern=r"[^@]+@[^@]+\.[^@]+", string=email):
        messagebox.showerror(title="Error", message="Email no válido")
        return
    # Guardar los datos en el archivo de Excel
    ws.append([nombre, apellido, edad, telefono, email])
    wb.save(nombre_archivo)
    # Limpiar los campos
    entry_nombre.delete(0, tk.END)  
    entry_apellido.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    # Mostrar mensaje de éxito  
    messagebox.showinfo(title="Éxito", message="Datos guardados correctamente")




root = tk.Tk()
root.title("Formulario de Registro")
root.configure(bg='#4B6587')
label_style = {
    "bg": "#4B6587",
    "fg": "white",
    "font": ("Arial", 12)
}
entry_style = {
    "bg": "#D3D3D3",
    "fg": "black",
    "font": ("Arial", 12)
}

label_nombre = tk.Label(root, text="Nombre", **label_style)
label_nombre.grid(row=0, column=0, padx=10, pady=5)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady=5)


label_apellido = tk.Label(root, text="Apellido", **label_style)
label_apellido.grid(row=1, column=0, padx=10, pady=5)       
entry_apellido = tk.Entry(root, **entry_style)
entry_apellido.grid(row=1, column=1, padx=10, pady=5)

label_edad = tk.Label(root, text="Edad", **label_style)
label_edad.grid(row=2, column=0, padx=10, pady=5)
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=2, column=1, padx=10, pady=5)

label_telefono = tk.Label(root, text="Telefono", **label_style)
label_telefono.grid(row=3, column=0, padx=10, pady=5)
entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=3, column=1, padx=10, pady=5)

label_email = tk.Label(root, text="Email", **label_style)
label_email.grid(row=4, column=0, padx=10, pady=5)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=4, column=1, padx=10, pady=5)

boton_guardar = tk.Button(root, text="Guardar", command=guardar_datos, bg="#4B6587", fg="white", width=20 ,font=("Arial", 12))
boton_guardar.grid(row=6, column=0, columnspan=2, pady=10)
root.mainloop()